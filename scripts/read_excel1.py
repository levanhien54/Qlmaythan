# -*- coding: utf-8 -*-
import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'D:\QL may than\Bảng tính không có tiêu đề.xlsx')
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Dimensions: {ws.dimensions}')
    for row in ws.iter_rows(min_row=1, max_row=min(60, ws.max_row), values_only=True):
        print(list(row))
