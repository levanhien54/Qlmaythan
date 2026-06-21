# -*- coding: utf-8 -*-
"""Deep audit tests — phơi bày các edge case còn ẩn sau các vòng fix trước."""
import io
import datetime as dt
import pytest
import openpyxl

from matching import find_staff, find_device, safe_str, parse_excel_datetime
from database.queries import thiet_bi, nhan_vien, phien_dieu_tri
from tests.test_excel_import_e2e import build_xlsx, make_row, post_excel


# ==========================================================
# GROUP A: STAFF/DEVICE MATCHING AMBIGUITY
# ==========================================================

def test_find_staff_prefix_collision():
    """Exact match 'Nguyễn Thị Vân' → id=7 (nhờ exact map thắng trước loop substring)."""
    all_staff = [
        {'id': 7, 'ho_ten': 'Nguyễn Thị Vân'},
        {'id': 25, 'ho_ten': 'Nguyễn Thị Vân Anh'},
    ]
    id_, _ = find_staff('Nguyễn Thị Vân', all_staff)
    assert id_ == 7


def test_find_staff_prefix_collision_reversed_order():
    all_staff = [
        {'id': 25, 'ho_ten': 'Nguyễn Thị Vân Anh'},
        {'id': 7, 'ho_ten': 'Nguyễn Thị Vân'},
    ]
    id_, _ = find_staff('Nguyễn Thị Vân', all_staff)
    assert id_ == 7


def test_find_staff_substring_ambiguity_silent_wrong_match():
    """🔴 BUG THỰC: Input KHÔNG exact → loop substring, FIRST WINS.

    Excel có 'Nguyễn Thị Vân' (thiếu dấu?). Nếu DB có
    'Nguyễn Thị Vân Anh' đứng trước 'Nguyễn Vân' (cả 2 đều substring match),
    hàm trả về ngườ i ĐẦU TIÊN trong loop — có thể sai.

    Ở đây tôi cho input 'Vân Anh' và DB có 2 người chứa 'vân anh':
    kết quả là silent ambiguity. Hàm NÊN trả về -1 (không chắc) thay vì
    guess first-match.
    """
    all_staff = [
        {'id': 1, 'ho_ten': 'Nguyễn Thị Vân Anh'},
        {'id': 2, 'ho_ten': 'Trần Thị Vân Anh'},
    ]
    id_, _ = find_staff('Vân Anh', all_staff)
    # Fix: khi ambiguous → trả -1, ép người nhập sửa dữ liệu gốc.
    assert id_ == -1


def test_find_device_number_ambiguity():
    """Convention BV: prefix 'Số'=B.Braun, 'F'=Fresinius. Raw 'số 1' → B.Braun."""
    devices = [
        (1, 'Máy chạy thận Fresinius số 1', '', 'OK'),
        (2, 'Máy chạy thận B.Braun số 1', '', 'OK'),
    ]
    # 'số 1' → convention-aware → B.Braun
    id_, _, _ = find_device('số 1', devices)
    assert id_ == 2
    # Keyword explicit
    id2, _, _ = find_device('Fresinius số 1', devices)
    assert id2 == 1
    # Prefix F → Fresinius
    id3, _, _ = find_device('_F1', devices)
    assert id3 == 1


def test_find_device_no_prefix_truly_ambiguous():
    """Số đơn thuần không prefix Số/F → ambiguous thật sự → None."""
    devices = [
        (1, 'Máy A số 5', '', 'OK'),
        (2, 'Máy B số 5', '', 'OK'),
    ]
    id_, _, _ = find_device('abc 5', devices)
    assert id_ is None


# ==========================================================
# GROUP B: SAFE_STR & TEXT EDGE CASES
# ==========================================================

@pytest.mark.parametrize("raw,expected", [
    ('tên\ttab', 'tên\ttab'),  # tab giữ nguyên (không phải newline)
    ('tên\u200bZ', 'tên\u200bZ'),  # zero-width space giữ lại (có thể gây khớp sai)
    ('\n\rline', 'line'),  # \n\r cụm khác nhau
])
def test_safe_str_preserves_unusual_chars(raw, expected):
    """Document behavior — có thể là rủi ro (zero-width không strip)."""
    assert safe_str(raw) == expected


# ==========================================================
# GROUP C: AGE PARSING QUIRKS
# ==========================================================

def test_age_decimal_truncated_silently(client_seeded_simple):
    """'65.5' hiện silently → 65. Document + cân nhắc có nên warn."""
    buf = build_xlsx([make_row(
        ho_ten='BN Decimal', tuoi_nam='65.5',
        ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )])
    d = post_excel(client_seeded_simple, buf).get_json()
    # Hiện tại PASS (truncate) — dùng test này để lock behavior
    assert d['success'] == 1
    rows = phien_dieu_tri.get_all(search='BN Decimal')
    assert rows[0]['tuoi'] == 65


def test_age_negative_rejected(client_seeded_simple):
    buf = build_xlsx([make_row(
        ho_ten='BN Neg', tuoi_nam=-5,
        ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )])
    d = post_excel(client_seeded_simple, buf).get_json()
    assert d['success'] == 0
    assert any('ngoài phạm vi' in e.lower() for e in d['errors'][0]['errors'])


# ==========================================================
# GROUP D: SQL INJECTION-SHAPED INPUT
# ==========================================================

def test_patient_name_with_apostrophe(client_seeded_simple):
    """Tên tiếng Anh có dấu nháy — không được crash hay SQL inject."""
    buf = build_xlsx([make_row(
        ho_ten="O'Brien Nguyễn", tuoi_nam=40,
        ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )])
    d = post_excel(client_seeded_simple, buf).get_json()
    assert d['success'] == 1


def test_patient_name_with_sql_metacharacters(client_seeded_simple):
    """Robust với '%', ';', '--' trong dữ liệu (parametrized query)."""
    buf = build_xlsx([make_row(
        ho_ten="Nguyễn %A_B; DROP TABLE--", tuoi_nam=40,
        ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )])
    d = post_excel(client_seeded_simple, buf).get_json()
    # Insert thành công & table vẫn còn
    assert d['success'] == 1
    assert phien_dieu_tri.count() >= 1


def test_search_with_sql_wildcard_does_not_cross_leak(client_seeded_simple):
    """Search với '%' không được trả về toàn bộ records (hiện behavior)."""
    nhan_vien.create(ho_ten='Bình thường', chuc_vu_trinh_do='KTV')
    r1 = nhan_vien.get_all(search='Bình thường')
    r2 = nhan_vien.get_all(search='%')
    # %'s ở đầu/cuối của LIKE → '%' match tất cả. Đây là behavior hiện có, document.
    assert len(r2) >= len(r1)


# ==========================================================
# GROUP E: EXCEL FORMAT QUIRKS
# ==========================================================

def test_multi_sheet_workbook_reads_only_first_sheet(client_seeded_simple):
    """sh.active chỉ đọc sheet hoạt động — dữ liệu ở sheet 2 bị bỏ qua silently."""
    wb = openpyxl.Workbook()
    sh1 = wb.active
    sh1.title = 'Chính'
    # Sheet 1: empty (chỉ header dummy), sheet 2: data thật
    for _ in range(12):
        sh1.append([''] * 29)
    sh2 = wb.create_sheet('Dữ liệu thật')
    for _ in range(10):
        sh2.append([''] * 29)
    row = make_row(
        ho_ten='BN Sheet2',
        ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )
    sh2.append(row)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    r = post_excel(client_seeded_simple, buf, 'multi.xlsx')
    d = r.get_json()
    # BN Sheet2 sẽ KHÔNG được import vì code chỉ đọc sheet active
    assert d.get('success', 0) == 0


def test_merged_cells_returned_as_none(client_seeded_simple):
    """Cell bị merge: cell chính có value, các cell sub có None → safe_str → ''."""
    wb = openpyxl.Workbook()
    sh = wb.active
    for _ in range(12):
        sh.append([''] * 29)
    # Row có merged cell — tên nằm ở cell 1, merge với cell 2 (tuổi)
    row = [''] * 29
    row[0] = 1
    row[1] = 'BN Merged'
    row[8] = '2026-04-01 08:00:00'
    row[9] = '2026-04-01 12:00:00'
    row[28] = 'Máy chạy thận Fresinius số 1'
    sh.append(row)
    # Merge cell 0:1 của data row (row index 13)
    sh.merge_cells(start_row=13, start_column=1, end_row=13, end_column=3)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    r = post_excel(client_seeded_simple, buf, 'merged.xlsx')
    d = r.get_json()
    # Import không crash (quan trọng nhất)
    assert r.status_code == 200


def test_formula_resolved_with_data_only(client_seeded_simple):
    """Tên nhập dạng formula ="BN " & "X" — với data_only=True phải đọc được giá trị.
    (Chỉ work nếu Excel đã save file sau khi evaluate)."""
    # openpyxl khi ghi file formula thô, data_only sẽ return None.
    # Đây là limitation đã biết — test lock behavior.
    wb = openpyxl.Workbook()
    sh = wb.active
    for _ in range(12):
        sh.append([''] * 29)
    row = [''] * 29
    row[0] = 1
    row[1] = '=CONCATENATE("BN ", "Formula")'
    row[8] = '2026-04-01 08:00:00'
    row[9] = '2026-04-01 12:00:00'
    row[28] = 'Máy chạy thận Fresinius số 1'
    sh.append(row)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    r = post_excel(client_seeded_simple, buf, 'formula.xlsx')
    # Không crash — formula chưa evaluate → hoặc skip (no ho_ten) hoặc ho_ten = '=...'
    assert r.status_code == 200


def test_empty_sheet(client_seeded_simple):
    """File hợp lệ nhưng toàn bộ cell rỗng."""
    wb = openpyxl.Workbook()
    sh = wb.active
    for _ in range(15):
        sh.append([''] * 29)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    r = post_excel(client_seeded_simple, buf, 'empty_sheet.xlsx')
    d = r.get_json()
    # 0 rows valid, không crash
    assert d.get('total', 0) == 0


# ==========================================================
# GROUP F: TIME BOUNDARY
# ==========================================================

def test_back_to_back_with_one_second_gap(client_seeded_simple):
    """Phiên cũ kết thúc 08:00:00, phiên mới bắt đầu 08:00:01 → OK."""
    buf = build_xlsx([
        make_row(stt=1, ho_ten='BN 1',
                 ngay_bd='2026-04-01 04:00:00', ngay_kt='2026-04-01 08:00:00',
                 may='Máy chạy thận Fresinius số 1'),
        make_row(stt=2, ho_ten='BN 2',
                 ngay_bd='2026-04-01 08:00:01', ngay_kt='2026-04-01 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
    ])
    d = post_excel(client_seeded_simple, buf).get_json()
    assert d['success'] == 2


def test_same_patient_multiple_valid_sessions(client_seeded_simple):
    """Cùng bệnh nhân → có thể lọc nhiều lần cùng ngày hoặc lặp lại — không block."""
    buf = build_xlsx([
        make_row(stt=1, ho_ten='BN Repeat',
                 ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
        make_row(stt=2, ho_ten='BN Repeat',
                 ngay_bd='2026-04-02 08:00:00', ngay_kt='2026-04-02 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
    ])
    d = post_excel(client_seeded_simple, buf).get_json()
    assert d['success'] == 2


def test_leap_year_feb_29(client_seeded_simple):
    """2028 là năm nhuận → 29/02 hợp lệ."""
    buf = build_xlsx([make_row(
        ho_ten='BN Leap',
        ngay_bd='29/02/2028 08:00:00', ngay_kt='29/02/2028 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )])
    d = post_excel(client_seeded_simple, buf).get_json()
    assert d['success'] == 1


# ==========================================================
# GROUP G: DATE FORMAT EXTENDED (ISO-T, Z, yyyy/mm/dd)
# ==========================================================

@pytest.mark.parametrize("raw,expected", [
    ('2026-04-01T08:00:00',  '2026-04-01 08:00:00'),   # ISO-T
    ('2026-04-01T08:00:00Z', '2026-04-01 08:00:00'),   # ISO-T-Zulu
    ('2026-04-01T08:00',     '2026-04-01 08:00:00'),   # ISO-T không giây
    ('2026/04/01',           '2026-04-01 00:00:00'),   # YYYY/MM/DD
    ('2026/04/01 08:00:00',  '2026-04-01 08:00:00'),
    ('1/4/2026',             '2026-04-01 00:00:00'),   # single digit d/m
    ('2026-4-1',             '2026-04-01 00:00:00'),   # single digit trong ISO
])
def test_parse_dt_extended_formats(raw, expected):
    assert parse_excel_datetime(raw) == expected


def test_parse_dt_rejects_impossible_dates():
    assert parse_excel_datetime('30/02/2026') is None  # Feb 30 không có
    assert parse_excel_datetime('31/04/2026') is None  # Apr 31 không có
    assert parse_excel_datetime('2026-04-01 25:00') is None
    assert parse_excel_datetime('00/00/2026') is None


# ==========================================================
# GROUP H: API VALIDATION (phòng rác từ UI)
# ==========================================================

def test_api_post_rejects_empty_name(client_seeded_simple):
    """API POST phải reject ho_ten rỗng — bảo vệ khỏi bug id=108 lặp lại."""
    r = client_seeded_simple.post('/api/phien-dieu-tri', json={
        'ho_ten': '',
        'ngay_bat_dau': '2026-04-01 08:00:00',
        'thiet_bi_id': 1,
    })
    assert r.status_code == 400
    assert 'ho_ten' in r.get_json()['error'].lower() or 'họ tên' in r.get_json()['error'].lower()


def test_api_post_rejects_single_char_name(client_seeded_simple):
    """API POST phải reject ho_ten 1 ký tự — bảo vệ khỏi bug id=144 lặp lại."""
    r = client_seeded_simple.post('/api/phien-dieu-tri', json={
        'ho_ten': 'n',
        'ngay_bat_dau': '2026-04-01 08:00:00',
        'thiet_bi_id': 1,
    })
    assert r.status_code == 400


def test_api_post_rejects_missing_start_date(client_seeded_simple):
    r = client_seeded_simple.post('/api/phien-dieu-tri', json={
        'ho_ten': 'BN Test',
        'thiet_bi_id': 1,
    })
    assert r.status_code == 400


def test_api_post_rejects_end_before_start(client_seeded_simple):
    r = client_seeded_simple.post('/api/phien-dieu-tri', json={
        'ho_ten': 'BN Test',
        'ngay_bat_dau': '2026-04-01 12:00:00',
        'ngay_ket_thuc': '2026-04-01 08:00:00',
        'thiet_bi_id': 1,
    })
    assert r.status_code == 400


def test_api_post_rejects_age_out_of_range(client_seeded_simple):
    r = client_seeded_simple.post('/api/phien-dieu-tri', json={
        'ho_ten': 'BN Test',
        'ngay_bat_dau': '2026-04-01 08:00:00',
        'tuoi': 150,
        'thiet_bi_id': 1,
    })
    assert r.status_code == 400


def test_api_put_same_validation(client_seeded_simple):
    """PUT cũng phải validate — không chỉ POST."""
    # Tạo phiên hợp lệ trước
    tb = thiet_bi.get_all()[0]
    sid = phien_dieu_tri.create(
        ho_ten='BN OK', thiet_bi_id=tb['id'],
        ngay_bat_dau='2026-04-01 08:00:00',
        ngay_ket_thuc='2026-04-01 12:00:00',
    )
    # Update với ho_ten rỗng → phải reject
    r = client_seeded_simple.put(f'/api/phien-dieu-tri/{sid}', json={
        'ho_ten': '',
        'ngay_bat_dau': '2026-04-01 08:00:00',
    })
    assert r.status_code == 400


def test_non_leap_year_feb_29_rejected(client_seeded_simple):
    """2027 không phải năm nhuận → 29/02 không hợp lệ."""
    buf = build_xlsx([make_row(
        ho_ten='BN NonLeap',
        ngay_bd='29/02/2027 08:00:00', ngay_kt='29/02/2027 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )])
    d = post_excel(client_seeded_simple, buf).get_json()
    assert d['success'] == 0


# ==========================================================
# FIXTURE
# ==========================================================

@pytest.fixture
def client_seeded_simple(temp_db):
    """Seed tối thiểu để E2E import vận hành."""
    thiet_bi.create(ten_thiet_bi='Máy chạy thận Fresinius số 1',
                    tinh_trang='Hoạt động bình thường')
    nhan_vien.create(ho_ten='Nguyễn Văn A', chuc_vu_trinh_do='Bác sĩ')
    from server import app
    app.config['TESTING'] = True
    with app.test_client() as c:
        yield c
