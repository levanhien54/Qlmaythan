# -*- coding: utf-8 -*-
"""Tests với 2 file Excel THỰC trong thư mục dự án:
  - 011.3.xls — Phiên điều trị (103 rows, có 6 cặp trùng khung giờ)
  - Bảng tính không có tiêu đề.xlsx — HIỆN TẠI là danh sách nhân viên (bug cấu hình)

Test này lock behavior thực tế để tránh regression khi refactor import.
"""
import os
import pytest
import openpyxl
from database.queries import thiet_bi, nhan_vien


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLS_SESSIONS = os.path.join(PROJECT_ROOT, '011.3.xls')
XLSX_STAFF   = os.path.join(PROJECT_ROOT, 'Bảng tính không có tiêu đề.xlsx')

# Các file Excel THẬT chứa dữ liệu bệnh nhân → gitignored, KHÔNG có trên CI.
# Bỏ qua cả module khi thiếu file (cục bộ có file thì vẫn chạy đầy đủ).
pytestmark = pytest.mark.skipif(
    not (os.path.exists(XLS_SESSIONS) and os.path.exists(XLSX_STAFF)),
    reason="Cần file Excel thật (gitignored) — không có trên CI",
)


# ---------- File existence ----------

def test_real_excel_files_exist():
    assert os.path.exists(XLS_SESSIONS), f'Missing: {XLS_SESSIONS}'
    assert os.path.exists(XLSX_STAFF), f'Missing: {XLSX_STAFF}'


# ---------- Config mismatch khẳng định ----------

def test_staff_xlsx_is_actually_staff_not_devices():
    """REGRESSION LOCK: 'Bảng tính không có tiêu đề.xlsx' chỉ có 3 cột (STT/Họ tên/Chức vụ).
    Nếu có người đổi file này thành danh sách thiết bị (8 cột như schema),
    test này sẽ fail → buộc phải xem lại config.EXCEL_THIET_BI.
    """
    wb = openpyxl.load_workbook(XLSX_STAFF, data_only=True)
    sh = wb.active
    header = list(sh.iter_rows(max_row=1, values_only=True))[0]
    non_empty = [c for c in header if c]
    assert len(non_empty) == 3, (
        f'File chỉ nên có 3 cột (staff), có {len(non_empty)}: {non_empty}'
    )
    assert any('HỌ VÀ TÊN' in (c or '').upper() for c in header)
    # Đếm nhân viên (bỏ header + 1 empty row)
    rows = [r for r in sh.iter_rows(min_row=3, values_only=True) if r[0] is not None]
    assert len(rows) == 31, f'Expected 31 staff, got {len(rows)}'


# ---------- Import 011.3.xls qua Flask test_client ----------

@pytest.fixture
def client_seeded(temp_db):
    """Seed máy + nhân viên theo đúng tên trong DB production để match chính xác."""
    # 12 Fresinius → match "_F1" → "_F12" trong Excel
    for n in range(1, 13):
        thiet_bi.create(
            ten_thiet_bi=f'Máy chạy thận Fresinius số {n}',
            tinh_trang='Hoạt động bình thường',
        )
    # B.Braun — bao trùm các số xuất hiện trong 011.3.xls
    # (file dùng: Số 22, 23, 25, 26, 27, 30, 31, 32, 35, 40-51, 14, 16, 19, 29, 36, 9, 45, 48, 49, 50, 51)
    for n in list(range(8, 52)):
        thiet_bi.create(
            ten_thiet_bi=f'Máy chạy thận B.Braun số {n}',
            tinh_trang='Hoạt động bình thường',
        )
    # HDF Online số 2, 3
    for n in (1, 2, 3):
        thiet_bi.create(
            ten_thiet_bi=f'Máy lọc máu liên tục HDF Online số {n}',
            tinh_trang='Hoạt động bình thường',
        )
    # NIPRO số 5 — match "_Nip 5"
    thiet_bi.create(ten_thiet_bi='Máy chạy thận NIPRO SỐ 5',
                    tinh_trang='Hoạt động bình thường')

    # Seed staff từ file "Bảng tính không có tiêu đề.xlsx"
    wb = openpyxl.load_workbook(XLSX_STAFF, data_only=True)
    sh = wb.active
    for r in sh.iter_rows(min_row=3, values_only=True):
        if r[0] is None:
            continue
        nhan_vien.create(ho_ten=str(r[1]).strip(),
                         chuc_vu_trinh_do=str(r[2] or '').strip())

    from server import app
    app.config['TESTING'] = True
    with app.test_client() as c:
        yield c


def test_real_sessions_xls_imports(client_seeded):
    """File 011.3.xls: 103 data rows.

    Sau fix 'nip' keyword: 2 row '_Nip 5' match đúng NIPRO (không còn đi nhầm vào
    Fresinius 5 gây false overlap) → skipped giảm từ 6 → 4.
    4 overlap còn lại là trùng khung giờ THẬT trong nguồn Excel.
    """
    with open(XLS_SESSIONS, 'rb') as f:
        r = client_seeded.post(
            '/api/phien-dieu-tri/import-excel',
            data={'file': (f, '011.3.xls')},
            content_type='multipart/form-data',
        )
    assert r.status_code == 200
    d = r.get_json()
    assert d['ok'] is True
    assert d['total'] == 103
    assert d['skipped'] == 2
    assert d['success'] == 101
    # Tất cả 6 lỗi phải là trùng khung giờ (không phải lỗi data khác)
    for err in d['errors']:
        msgs = err['errors']
        assert any('Trùng' in m for m in msgs), (
            f'Row {err["row"]}: expected trùng giờ, got {msgs}'
        )


def test_disaster_recovery_import_thiet_bi_skips_staff_file(temp_db, monkeypatch, capsys):
    """B3 NEW FIX: nếu ai đó chạy import_thiet_bi() khi DB trống VÀ
    file EXCEL_THIET_BI là danh sách nhân viên (config mismatch hiện tại),
    import phải REFUSE thay vì tạo 31 'thiết bị' sai."""
    import config
    monkeypatch.setattr(config, 'EXCEL_THIET_BI', XLSX_STAFF)
    # Đảm bảo import_data dùng path mới
    import import_data
    monkeypatch.setattr(import_data, 'EXCEL_THIET_BI', XLSX_STAFF)

    import_data.import_thiet_bi()
    assert thiet_bi.count() == 0, 'Phải skip — không tạo thiết bị từ file staff'
    out = capsys.readouterr().out
    assert 'KHÔNG CÓ cấu trúc' in out or 'skip' in out.lower()


def test_real_import_idempotent_second_run_blocks_all(client_seeded):
    """Import 011.3.xls lần 2 ngay sau lần 1 → 0 success, 97 trùng, 6 trùng = 103 skipped.

    Đây là regression guard cho BUG PRODUCTION THỰC: data đã bị
    duplicate khi import 2 lần trước khi B2 được fix.
    """
    with open(XLS_SESSIONS, 'rb') as f:
        client_seeded.post(
            '/api/phien-dieu-tri/import-excel',
            data={'file': (f, '011.3.xls')},
            content_type='multipart/form-data',
        )
    with open(XLS_SESSIONS, 'rb') as f:
        r = client_seeded.post(
            '/api/phien-dieu-tri/import-excel',
            data={'file': (f, '011.3.xls')},
            content_type='multipart/form-data',
        )
    d = r.get_json()
    assert d['success'] == 0, f'Import lần 2 KHÔNG ĐƯỢC tạo phiên mới! (got {d["success"]})'
    assert d['skipped'] == 103
