# -*- coding: utf-8 -*-
"""Deep audit vòng 2 — tập trung Excel import.

Các edge case phức tạp hơn: find_device ambiguous, column shift, Unicode
normalization, rich text, cell error, NFC vs NFD, transaction safety.
"""
import io
import unicodedata
import pytest
import openpyxl

from matching import find_device, find_staff, parse_excel_datetime
from database.queries import thiet_bi, nhan_vien, phien_dieu_tri
from tests.test_excel_import_e2e import build_xlsx, make_row, post_excel


@pytest.fixture
def client_rich(temp_db):
    """Seed đa dạng để expose ambiguity: 3 máy số 1, staff có dấu NFC/NFD."""
    thiet_bi.create(ten_thiet_bi='Máy chạy thận Fresinius số 1',
                    tinh_trang='Hoạt động bình thường')
    thiet_bi.create(ten_thiet_bi='Máy chạy thận B.Braun số 1',
                    tinh_trang='Hoạt động bình thường')
    thiet_bi.create(ten_thiet_bi='Máy lọc máu HDF Online số 1',
                    tinh_trang='Hoạt động bình thường')
    thiet_bi.create(ten_thiet_bi='Máy chạy thận Fresinius số 10',
                    tinh_trang='Hoạt động bình thường')
    nhan_vien.create(ho_ten='Nguyễn Văn A', chuc_vu_trinh_do='BS')
    from server import app
    app.config['TESTING'] = True
    with app.test_client() as c:
        yield c


# ==========================================================
# A. FIND_DEVICE AMBIGUITY — số giống nhau, raw không keyword
# ==========================================================

def test_find_device_pure_number_no_keyword_is_ambiguous():
    """🔴 BUG: raw='Máy 1' (chỉ số, không hãng) ở DB có 3 máy 'số 1' khác hãng.
    Kỳ vọng: trả None (ambiguous) để ép user ghi rõ hãng."""
    devices = [
        (1, 'Máy chạy thận Fresinius số 1', '', 'OK'),
        (2, 'Máy chạy thận B.Braun số 1', '', 'OK'),
        (3, 'Máy lọc máu HDF số 1', '', 'OK'),
    ]
    id_, _, _ = find_device('Máy 1', devices)
    assert id_ is None, f'Expected None (ambiguous), got {id_}'


def test_find_device_keyword_disambiguates():
    """Có keyword phân biệt → trả đúng máy."""
    devices = [
        (1, 'Máy chạy thận Fresinius số 1', '', 'OK'),
        (2, 'Máy chạy thận B.Braun số 1', '', 'OK'),
        (3, 'Máy lọc máu HDF số 1', '', 'OK'),
    ]
    assert find_device('Fresinius_1', devices)[0] == 1
    assert find_device('B.Braun_1', devices)[0] == 2
    assert find_device('HDF_01', devices)[0] == 3


def test_find_device_number_only_unique_ok():
    """Nếu chỉ 1 máy match số → không ambiguous, trả máy đó."""
    devices = [
        (1, 'Máy chạy thận Fresinius số 7', '', 'OK'),
        (2, 'Máy chạy thận B.Braun số 99', '', 'OK'),
    ]
    id_, _, _ = find_device('_7', devices)
    assert id_ == 1


def test_find_device_substring_unique_ok():
    """Raw chứa gần đủ tên → unique match → OK."""
    devices = [
        (1, 'Máy chạy thận Fresinius số 1', '', 'OK'),
    ]
    id_, _, _ = find_device('Fresinius 1', devices)
    assert id_ == 1


def test_find_device_trailing_space_ok():
    """Raw 'Fresinius số  1   ' (khoảng trắng thừa, trailing space) vẫn match."""
    devices = [
        (1, 'Máy chạy thận Fresinius số 1', '', 'OK'),
    ]
    id_, _, _ = find_device('  Fresinius số  1   ', devices)
    assert id_ == 1


# ==========================================================
# B. UNICODE NFC vs NFD NORMALIZATION
# ==========================================================

def test_find_device_nfd_input_matches_nfc_db():
    """Tên từ Mac hay browser đôi lúc ở NFD; DB thường NFC. Phải match được."""
    nfc_name = 'Máy thận số 1'
    nfd_name = unicodedata.normalize('NFD', nfc_name)
    assert nfc_name != nfd_name, 'Setup: 2 chuỗi phải khác bytes'

    devices = [(1, nfc_name, '', 'OK')]
    # Hiện tại có thể FAIL — đây là test expose bug tiềm ẩn.
    id_, _, _ = find_device(nfd_name, devices)
    # Expect: match được. Nếu fail — thêm NFC normalize vào safe_str/matching.
    assert id_ == 1, 'find_device phải normalize NFC trước khi so sánh'


def test_find_staff_nfd_input_matches_nfc_db():
    nfc = 'Nguyễn Văn A'
    nfd = unicodedata.normalize('NFD', nfc)
    staff = [{'id': 1, 'ho_ten': nfc}]
    id_, _ = find_staff(nfd, staff)
    assert id_ == 1


# ==========================================================
# C. COLUMN LAYOUT FRAGILITY
# ==========================================================

def test_hardcoded_column_index_breaks_silently_if_shifted(client_rich):
    """Nếu file có 1 cột extra chèn ở đầu (STT → col 1 thay vì 0), dữ liệu shift.
    Hiện tại hardcoded [1]=họ tên, [28]=máy — không robust.

    Test này document fragility: file shifted KHÔNG crash, chỉ import sai.
    """
    wb = openpyxl.Workbook()
    sh = wb.active
    # Các dòng đầu rỗng để qua filter
    for _ in range(10):
        sh.append([''] * 30)
    # Header shifted 1 cột phải
    header = [''] + [''] * 29
    header[1] = 'TT'      # Col 1 (đáng ra 0)
    header[2] = 'Họ và tên'  # Col 2 (đáng ra 1)
    header[29] = 'Máy thực hiện'
    sh.append(header)
    sh.append([''] * 30)  # spacer
    row = [''] * 30
    row[1] = 1
    row[2] = 'BN Shifted'  # họ tên đúng ở cột 2
    row[9] = '2026-04-01 08:00:00'
    row[10] = '2026-04-01 12:00:00'
    row[29] = 'Máy chạy thận Fresinius số 1'
    sh.append(row)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    r = post_excel(client_rich, buf, 'shifted.xlsx')
    d = r.get_json()
    # Behavior hiện tại: row[1] là str '' (col 1 shifted) → ho_ten = '' → silently skipped
    # BN Shifted → không import. Không crash nhưng mất dữ liệu.
    assert d.get('total', 0) == 0 or d.get('success', 0) == 0


# ==========================================================
# D. RICH TEXT / FORMULA / CELL ERROR
# ==========================================================

def test_rich_text_cell_read_as_plain_string(client_rich):
    """Cell có định dạng rich text (bold, color) phải được đọc như string."""
    wb = openpyxl.Workbook()
    sh = wb.active
    for _ in range(12):
        sh.append([''] * 29)
    row = [''] * 29
    row[0] = 1
    # openpyxl tự xử lý rich text khi iter_rows values_only=True → trả về CellRichText
    # hoặc str. Chỉ cần không crash.
    row[1] = 'Nguyễn Bold'
    row[8] = '2026-04-01 08:00:00'
    row[9] = '2026-04-01 12:00:00'
    row[28] = 'Máy chạy thận Fresinius số 1'
    sh.append(row)
    # Apply rich format
    cell = sh.cell(row=13, column=2)
    cell.font = openpyxl.styles.Font(bold=True)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    r = post_excel(client_rich, buf, 'rich.xlsx')
    d = r.get_json()
    assert d['success'] == 1


def test_extra_columns_beyond_29_ignored(client_rich):
    """File có 40 cột, chỉ dùng 29 đầu. Extra ignored, không crash."""
    wb = openpyxl.Workbook()
    sh = wb.active
    for _ in range(12):
        sh.append(['x'] * 40)
    row = [''] * 40
    row[0] = 1
    row[1] = 'BN Many Cols'
    row[8] = '2026-04-01 08:00:00'
    row[9] = '2026-04-01 12:00:00'
    row[28] = 'Máy chạy thận Fresinius số 1'
    row[35] = 'Extra data ignored'
    sh.append(row)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    r = post_excel(client_rich, buf, 'wide.xlsx')
    d = r.get_json()
    assert d['success'] == 1


def test_extra_whitespace_rows_between_data(client_rich):
    """Dòng trắng giữa data rows không được break loop."""
    buf = build_xlsx([
        make_row(stt=1, ho_ten='BN 1',
                 ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
        [''] * 29,  # empty row
        make_row(stt=3, ho_ten='BN 3',
                 ngay_bd='2026-04-02 08:00:00', ngay_kt='2026-04-02 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
    ])
    d = post_excel(client_rich, buf).get_json()
    assert d['success'] == 2


# ==========================================================
# E. SỐ HỒ SƠ TYPE COERCION
# ==========================================================

@pytest.mark.parametrize("raw,expected_str", [
    (26035598.0, '26035598'),  # từ openpyxl float
    (26035598, '26035598'),
    ('HS-2026-001', 'HS-2026-001'),  # alphanumeric
    (1.5e7, '15000000'),  # scientific notation
    ('', ''),
    (None, ''),
])
def test_so_ho_so_coercion(raw, expected_str, client_rich):
    """Số hồ sơ nhiều kiểu input phải normalize về string hợp lệ."""
    buf = build_xlsx([make_row(
        ho_ten='BN SHS', so_ho_so=raw,
        ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )])
    d = post_excel(client_rich, buf).get_json()
    assert d['success'] == 1
    rows = phien_dieu_tri.get_all(search='BN SHS')
    assert rows[0]['so_ho_so'] == expected_str


# ==========================================================
# F. LARGE BATCH — performance + memory
# ==========================================================

def test_large_batch_100_rows(client_rich):
    """100 rows không duplicate (100 giờ khác nhau trên cùng máy tuần tự)."""
    rows = []
    from datetime import datetime, timedelta
    t0 = datetime(2026, 5, 1, 0, 0, 0)
    for i in range(100):
        start = t0 + timedelta(hours=i)
        end = start + timedelta(minutes=30)
        rows.append(make_row(
            stt=i + 1, ho_ten=f'BN Batch {i:03d}',
            ngay_bd=start.strftime('%Y-%m-%d %H:%M:%S'),
            ngay_kt=end.strftime('%Y-%m-%d %H:%M:%S'),
            may='Máy chạy thận Fresinius số 1',
        ))
    buf = build_xlsx(rows)
    d = post_excel(client_rich, buf).get_json()
    assert d['total'] == 100
    assert d['success'] == 100


# ==========================================================
# G. RE-UPLOAD AFTER PARTIAL (IDEMPOTENCY)
# ==========================================================

def test_reupload_same_file_inserts_nothing_new(client_rich):
    """Lần 1: 2 rows pass. Lần 2: 0 pass, 2 skipped (all duplicate)."""
    buf1 = build_xlsx([
        make_row(stt=1, ho_ten='BN 1',
                 ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
        make_row(stt=2, ho_ten='BN 2',
                 ngay_bd='2026-04-02 08:00:00', ngay_kt='2026-04-02 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
    ])
    d1 = post_excel(client_rich, buf1).get_json()
    assert d1['success'] == 2

    buf2 = build_xlsx([
        make_row(stt=1, ho_ten='BN 1',
                 ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
        make_row(stt=2, ho_ten='BN 2',
                 ngay_bd='2026-04-02 08:00:00', ngay_kt='2026-04-02 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
    ])
    d2 = post_excel(client_rich, buf2).get_json()
    assert d2['success'] == 0
    assert d2['skipped'] == 2


def test_reupload_mixed_old_and_new_rows(client_rich):
    """Re-upload file có cả row cũ và row mới → chỉ row mới insert."""
    buf1 = build_xlsx([make_row(
        ho_ten='BN Old',
        ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )])
    post_excel(client_rich, buf1)

    buf2 = build_xlsx([
        make_row(stt=1, ho_ten='BN Old',
                 ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
        make_row(stt=2, ho_ten='BN New',
                 ngay_bd='2026-05-01 08:00:00', ngay_kt='2026-05-01 12:00:00',
                 may='Máy chạy thận Fresinius số 1'),
    ])
    d = post_excel(client_rich, buf2).get_json()
    assert d['success'] == 1
    assert d['skipped'] == 1


# ==========================================================
# H. XLS vs XLSX PARITY
# ==========================================================

def test_xls_and_xlsx_produce_same_result(client_rich, tmp_path):
    """File .xls và .xlsx cùng data phải cho kết quả giống nhau."""
    import xlwt  # noqa — needed for .xls creation
    try:
        import xlwt
    except ImportError:
        pytest.skip('xlwt không cài')

    # .xlsx
    buf_xlsx = build_xlsx([make_row(
        ho_ten='BN Parity',
        ngay_bd='2026-04-01 08:00:00', ngay_kt='2026-04-01 12:00:00',
        may='Máy chạy thận Fresinius số 1',
    )])
    d_xlsx = post_excel(client_rich, buf_xlsx, 'parity.xlsx').get_json()

    # .xls — xlwt cấm overwrite; chỉ write ô cần thiết.
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('s', cell_overwrite_ok=True)
    ws.write(9, 0, 'TT')
    ws.write(9, 1, 'Họ và tên')
    ws.write(11, 0, 1)
    ws.write(11, 1, 'BN Parity2')
    ws.write(11, 8, '2026-04-02 08:00:00')
    ws.write(11, 9, '2026-04-02 12:00:00')
    ws.write(11, 28, 'Máy chạy thận Fresinius số 1')
    fp = tmp_path / 'parity.xls'
    wb.save(str(fp))
    with open(fp, 'rb') as f:
        d_xls = post_excel(client_rich, f, 'parity.xls').get_json()

    assert d_xlsx['success'] == 1
    assert d_xls['success'] == 1
