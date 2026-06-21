# -*- coding: utf-8 -*-
"""
Import data from Excel files into SQLite database.
Run once or when data needs updating.
"""
import os
import re
import sys
import openpyxl
import xlrd
from database.models import create_all_tables
from database.connection import db
from database.queries import nhan_vien, thiet_bi
from config import EXCEL_THIET_BI, EXCEL_PHIEN_DT


def normalize_tinh_trang(raw: str) -> str:
    if not raw:
        return "Khong ro"
    raw = raw.strip().lower()
    if "hỏng" in raw:
        return "Hỏng"
    if "lỗi" in raw or "nghỉ" in raw:
        return "Báo lỗi"
    if "bình thường" in raw or "bt" in raw:
        return "Hoạt động bình thường"
    return raw.strip()


def parse_machine_name(raw: str) -> str:
    if not raw:
        return ""
    parts = raw.split("_")
    if len(parts) > 1:
        return parts[-1].strip()
    return raw.strip()


def parse_datetime_xls(value, wb_datemode):
    if isinstance(value, float) and value > 0:
        try:
            dt = xlrd.xldate_as_tuple(value, wb_datemode)
            return f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d} {dt[3]:02d}:{dt[4]:02d}:{dt[5]:02d}"
        except Exception:
            return None
    if isinstance(value, str) and value.strip():
        try:
            parts = value.strip().split(" ")
            date_parts = parts[0].split("/")
            if len(date_parts) == 3:
                d, m, y = date_parts
                time_part = parts[1] if len(parts) > 1 else "00:00:00"
                if len(time_part.split(":")) == 2:
                    time_part += ":00"
                # Zero-pad để ra ISO hợp lệ ('2026-3-5' sai thứ tự sắp xếp chuỗi).
                return f"{int(y):04d}-{int(m):02d}-{int(d):02d} {time_part}"
        except Exception:
            pass
    return None


def _looks_like_device_sheet(ws) -> bool:
    """Heuristic: file device phải có >=6 cột non-empty ở row 1 và header
    chứa từ khóa 'thiết bị' / 'model' / 'số máy'.
    Nếu file chỉ có 3 cột (STT / Họ tên / Chức vụ) → return False."""
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    non_empty = [c for c in row1 if c not in (None, '')]
    if len(non_empty) < 6:
        return False
    header_text = ' '.join(str(c).lower() for c in non_empty)
    return any(k in header_text for k in ('thiết bị', 'model', 'số máy', 'hãng sx', 'hãng sản xuất'))


def import_thiet_bi():
    print("[Import] Loading devices...")
    if thiet_bi.count() > 0:
        print("[Import] thiet_bi already has data, skipping.")
        return

    if not os.path.exists(EXCEL_THIET_BI):
        print(f"[Import] ⚠️  File không tồn tại: {EXCEL_THIET_BI} — skip.")
        return

    wb = openpyxl.load_workbook(EXCEL_THIET_BI)
    ws = wb.active

    if not _looks_like_device_sheet(ws):
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        print(f"[Import] ⚠️  File '{EXCEL_THIET_BI}' KHÔNG CÓ cấu trúc danh sách thiết bị.")
        print(f"[Import]    Header thực tế: {[c for c in row1 if c]}")
        print(f"[Import]    Expected: STT, Tên thiết bị, Model, Hãng SX, Số máy, Năm SD, Tình trạng, Tần suất")
        print(f"[Import]    → Bỏ qua để tránh hỏng dữ liệu. Sửa config.EXCEL_THIET_BI.")
        return

    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        stt, ten, model, hang_sx, so_may, nam, tinh_trang_raw, tan_suat = row[:8]
        if not ten:
            continue
        thiet_bi.create(
            ten_thiet_bi=str(ten).strip(),
            model=str(model or "").strip(),
            hang_san_xuat=str(hang_sx or "").strip(),
            nuoc_san_xuat=str(hang_sx or "").strip(),
            so_may=str(so_may or "").strip(),
            nam_su_dung=int(nam) if nam else 0,
            tinh_trang=normalize_tinh_trang(str(tinh_trang_raw or "")),
            tan_suat_su_dung=int(tan_suat) if tan_suat else 0,
        )
        count += 1
    print(f"[Import] Imported {count} devices.")


def import_phien_dieu_tri():
    print("[Import] Loading treatment sessions...")
    from database.queries import phien_dieu_tri as pdt_q
    if pdt_q.count() > 0:
        print("[Import] phien_dieu_tri already has data, skipping.")
        return

    wb = xlrd.open_workbook(EXCEL_PHIEN_DT)
    ws = wb.sheet_by_index(0)

    count = 0
    for r in range(11, ws.nrows):
        try:
            stt = ws.cell_value(r, 0)
            if not stt:
                continue
            ho_ten = str(ws.cell_value(r, 1)).strip()
            if not ho_ten:
                continue

            # Tuổi/Số hồ sơ: chuyển đổi KHOAN DUNG — giá trị không phải số không
            # được làm hỏng (raise) cả dòng phiên; chỉ bỏ qua riêng trường đó.
            def _to_int(v):
                try:
                    return int(float(str(v).strip()))
                except (ValueError, TypeError):
                    return 0

            tuoi_nam = ws.cell_value(r, 2)
            tuoi_nu = ws.cell_value(r, 3)
            tuoi = _to_int(tuoi_nam) if tuoi_nam else _to_int(tuoi_nu)

            dia_chi = str(ws.cell_value(r, 4) or "").strip()
            so_ho_so_raw = ws.cell_value(r, 5)
            if isinstance(so_ho_so_raw, (int, float)):
                so_ho_so = str(int(so_ho_so_raw))
            else:
                # Số hồ sơ chữ-số (vd 'HS-2024-01') giữ nguyên thay vì làm rớt dòng.
                so_ho_so = str(so_ho_so_raw or "").strip()

            ngay_bd = parse_datetime_xls(ws.cell_value(r, 8), wb.datemode)
            ngay_kt = parse_datetime_xls(ws.cell_value(r, 9), wb.datemode)

            ptv_chinh_ten = str(ws.cell_value(r, 18) or "").strip()
            phu_1_ten = str(ws.cell_value(r, 19) or "").strip()
            may_raw = str(ws.cell_value(r, 28) or "").strip()

            ptv_chinh_id = nhan_vien.get_or_create(ptv_chinh_ten, "Bac si") if ptv_chinh_ten else None
            phu_1_id = nhan_vien.get_or_create(phu_1_ten, "Dieu duong") if phu_1_ten else None

            machine_name = parse_machine_name(may_raw)
            tb = thiet_bi.find_by_name(machine_name) if machine_name else None

            pdt_q.create(
                ho_ten=ho_ten, may_thuc_hien=may_raw,
                thiet_bi_id=tb["id"] if tb else None,
                tuoi=tuoi, dia_chi=dia_chi, so_ho_so=so_ho_so,
                ngay_bat_dau=ngay_bd, ngay_ket_thuc=ngay_kt,
                ptv_chinh_id=ptv_chinh_id, phu_1_id=phu_1_id,
            )
            count += 1
        except Exception as e:
            print(f"[Import] Error row {r}: {e}")
            continue
    print(f"[Import] Imported {count} sessions.")


def run_import():
    print("=" * 50)
    print("STARTING DATA IMPORT")
    print("=" * 50)
    create_all_tables()
    import_thiet_bi()
    import_phien_dieu_tri()
    print("=" * 50)
    print(f"IMPORT COMPLETE: devices={thiet_bi.count()}, staff={nhan_vien.count()}")
    from database.queries import phien_dieu_tri as pdt_q
    print(f"                 sessions={pdt_q.count()}")
    print("=" * 50)


if __name__ == "__main__":
    run_import()
