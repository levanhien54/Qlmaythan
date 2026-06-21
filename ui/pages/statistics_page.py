# -*- coding: utf-8 -*-
"""
Trang Thống kê & Báo cáo.
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QScrollArea, QFrame, QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QPainter, QColor
from ui.styles import COLORS, btn_success
from database.queries import thiet_bi, phien_dieu_tri, bao_duong
from config import TAN_SUAT


class BarChartWidget(QFrame):
    """Biểu đồ bar ngang."""

    def __init__(self, title: str, parent=None):
        super().__init__(parent)
        self._title = title
        self._data = {}
        self.setMinimumHeight(350)
        c = COLORS
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 12px;
            }}
        """)

    def set_data(self, data: dict):
        self._data = data
        self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        c = COLORS

        w, h = self.width(), self.height()
        m = 20

        # Title
        painter.setPen(QColor(c["text_primary"]))
        painter.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        painter.drawText(m, m + 18, self._title)

        if not self._data:
            painter.setPen(QColor(c["text_muted"]))
            painter.setFont(QFont("Segoe UI", 12))
            painter.drawText(m, m + 50, "Chưa có dữ liệu")
            painter.end()
            return

        chart_x = m + 180
        chart_y = m + 40
        chart_w = w - chart_x - m
        chart_h = h - chart_y - m

        max_val = max(self._data.values()) if self._data.values() else 1
        bar_h = min(28, max(14, chart_h // max(len(self._data), 1) - 4))
        colors = ["#e94560", "#1a73e8", "#00c853", "#ffc107",
                  "#9c27b0", "#ff9800", "#00bcd4", "#8bc34a"]

        for i, (label, value) in enumerate(list(self._data.items())[:20]):
            y = chart_y + i * (bar_h + 4)
            if y + bar_h > h - m:
                break

            # Label
            painter.setPen(QColor(c["text_secondary"]))
            painter.setFont(QFont("Segoe UI", 9))
            label_text = str(label)[:25]
            painter.drawText(m, y + bar_h - 4, label_text)

            # Bar
            bar_w = int((value / max_val) * chart_w) if max_val > 0 else 0
            color = QColor(colors[i % len(colors)])
            painter.setBrush(color)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(chart_x, y, max(bar_w, 3), bar_h, 4, 4)

            # Value
            painter.setPen(QColor(c["text_primary"]))
            painter.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
            painter.drawText(chart_x + bar_w + 6, y + bar_h - 4, str(value))

        painter.end()


class StatisticsPage(QWidget):
    """Trang thống kê & báo cáo."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()

    def _setup_ui(self):
        c = COLORS

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        container = QWidget()
        container.setStyleSheet("background: transparent;")
        layout = QVBoxLayout(container)
        layout.setSpacing(20)
        layout.setContentsMargins(24, 24, 24, 24)

        # Header
        header = QHBoxLayout()
        title = QLabel("📈 Thống kê & Báo cáo")
        title.setFont(QFont("Segoe UI", 20, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {c['text_primary']}; background: transparent;")
        header.addWidget(title)
        header.addStretch()

        btn_export = QPushButton("📊 Xuất Excel")
        btn_export.setStyleSheet(btn_success())
        btn_export.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_export.clicked.connect(self._export)
        header.addWidget(btn_export)

        layout.addLayout(header)

        # Summary cards row
        summary_layout = QHBoxLayout()
        summary_layout.setSpacing(16)

        self.card_total_cost = self._make_summary_card("💰", "0 VNĐ", "Tổng chi phí BD")
        self.card_top_machine = self._make_summary_card("🏆", "-", "Máy dùng nhiều nhất")
        self.card_active_rate = self._make_summary_card("📊", "0%", "Tỷ lệ hoạt động")

        summary_layout.addWidget(self.card_total_cost)
        summary_layout.addWidget(self.card_top_machine)
        summary_layout.addWidget(self.card_active_rate)
        layout.addLayout(summary_layout)

        # Charts
        self.chart_usage = BarChartWidget("Tần suất sử dụng theo máy")
        self.chart_usage.setMinimumHeight(400)
        layout.addWidget(self.chart_usage)

        self.chart_sessions = BarChartWidget("Số phiên theo máy")
        self.chart_sessions.setMinimumHeight(400)
        layout.addWidget(self.chart_sessions)

        layout.addStretch()
        scroll.setWidget(container)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll)

    def _make_summary_card(self, icon: str, value: str, label: str) -> QFrame:
        c = COLORS
        frame = QFrame()
        frame.setFixedHeight(100)
        frame.setStyleSheet(f"""
            QFrame {{
                background-color: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 12px;
            }}
        """)
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(16, 12, 16, 12)

        top = QLabel(f"{icon} {label}")
        top.setFont(QFont("Segoe UI", 11))
        top.setStyleSheet(f"color: {c['text_secondary']}; background: transparent;")
        layout.addWidget(top)

        val_label = QLabel(value)
        val_label.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        val_label.setStyleSheet(f"color: {c['accent_blue_light']}; background: transparent;")
        layout.addWidget(val_label)

        frame._val_label = val_label
        return frame

    def refresh_data(self):
        # Total cost
        total_cost = bao_duong.total_chi_phi()
        self.card_total_cost._val_label.setText(f"{total_cost:,.0f} VNĐ")

        # Top machine
        sessions = phien_dieu_tri.sessions_per_machine()
        if sessions:
            top = sessions[0]
            self.card_top_machine._val_label.setText(
                f"{top['may_thuc_hien']} ({top['so_phien']} phiên)"
            )

        # Active rate
        total = thiet_bi.count()
        status = thiet_bi.count_by_tinh_trang()
        active = status.get("Hoạt động", 0)
        rate = int((active / total) * 100) if total > 0 else 0
        self.card_active_rate._val_label.setText(f"{rate}%")

        # Chart: Usage per device
        all_devices = thiet_bi.get_all()
        usage_data = {}
        for d in all_devices:
            name = d["ten_thiet_bi"]
            if len(name) > 30:
                name = name[:28] + "..."
            usage_data[name] = d.get("tan_suat_su_dung", 0)
        # Sort by usage desc
        usage_data = dict(sorted(usage_data.items(), key=lambda x: x[1], reverse=True))
        self.chart_usage.set_data(usage_data)

        # Chart: Sessions per machine
        sess_data = {s["may_thuc_hien"]: s["so_phien"] for s in sessions[:20]}
        self.chart_sessions.set_data(sess_data)

    def _export(self):
        """Xuất báo cáo ra Excel."""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Lưu báo cáo", "bao_cao_thiet_bi.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            import openpyxl
            wb = openpyxl.Workbook()

            # Sheet 1: Thiết bị
            ws1 = wb.active
            ws1.title = "Thiết bị"
            ws1.append(["STT", "Tên thiết bị", "Model", "Hãng SX",
                        "Số máy", "Năm SĐ", "Tình trạng", "Tần suất"])
            for i, d in enumerate(thiet_bi.get_all(), 1):
                ws1.append([
                    i, d["ten_thiet_bi"], d["model"], d["hang_san_xuat"],
                    d["so_may"], d["nam_su_dung"], d["tinh_trang"],
                    TAN_SUAT.get(d["tan_suat_su_dung"], "")
                ])

            # Sheet 2: Bảo dưỡng
            ws2 = wb.create_sheet("Bảo dưỡng")
            ws2.append(["STT", "Thiết bị", "Loại", "Ngày TH",
                        "Người TH", "Mô tả", "Chi phí", "Trạng thái"])
            for i, d in enumerate(bao_duong.get_all(), 1):
                ws2.append([
                    i, d.get("ten_thiet_bi", ""), d["loai"],
                    d.get("ngay_thuc_hien", ""),
                    d.get("nguoi_thuc_hien_ten", ""),
                    d.get("mo_ta", ""), d.get("chi_phi", 0),
                    d.get("trang_thai", "")
                ])

            # Sheet 3: Phiên ĐT
            ws3 = wb.create_sheet("Phiên điều trị")
            ws3.append(["STT", "Họ tên", "Tuổi", "Địa chỉ", "Số HS",
                        "Ngày BĐ", "Ngày KT", "PTV chính", "Phụ 1", "Máy"])
            for i, d in enumerate(phien_dieu_tri.get_all(), 1):
                ws3.append([
                    i, d["ho_ten"], d["tuoi"], d.get("dia_chi", ""),
                    d.get("so_ho_so", ""),
                    d.get("ngay_bat_dau", ""), d.get("ngay_ket_thuc", ""),
                    d.get("ptv_chinh_ten", ""), d.get("phu_1_ten", ""),
                    d.get("may_thuc_hien", "")
                ])

            wb.save(file_path)
            QMessageBox.information(self, "Thành công",
                                    f"Đã xuất báo cáo:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Lỗi xuất Excel: {e}")
